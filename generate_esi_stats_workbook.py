from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


DEFAULT_WORKDIR = Path(__file__).resolve().parent
DEFAULT_ANALYSIS = DEFAULT_WORKDIR / "analysis"
DEFAULT_OUTPUT = DEFAULT_WORKDIR / "outputs" / "2025年度ESI高水平论文统计结果.xlsx"
DEFAULT_DETAILED_OUTPUT = DEFAULT_WORKDIR / "outputs" / "2025年度ESI高水平论文去重总表.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据 analysis 结果生成统计工作簿")
    parser.add_argument("--analysis-dir", type=Path, default=DEFAULT_ANALYSIS, help="analysis 目录")
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT, help="输出 xlsx 路径")
    parser.add_argument("--detailed-output-xlsx", type=Path, default=DEFAULT_DETAILED_OUTPUT, help="详细总表 xlsx 路径")
    return parser.parse_args()


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def dedupe_preserve(items) -> list[str]:
    seen = set()
    ordered = []
    for item in items:
        if not item or item in seen:
            continue
        seen.add(item)
        ordered.append(item)
    return ordered


def apply_header(ws, row: int, start_col: int, end_col: int, fill: str = "1F4E78") -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_table_style(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def autofit_like(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_distribution_sheet(wb: Workbook, summary: dict) -> None:
    ws = wb.active
    ws.title = "表12_入围期数"
    ws.append(["类型", "入围1期", "入围2期", "入围3期", "入围4期", "入围5期", "入围6期", "合计"])
    ws.append(
        [
            "高被引论文",
            summary["distribution"]["high_cited"]["1"],
            summary["distribution"]["high_cited"]["2"],
            summary["distribution"]["high_cited"]["3"],
            summary["distribution"]["high_cited"]["4"],
            summary["distribution"]["high_cited"]["5"],
            summary["distribution"]["high_cited"]["6"],
            summary["totals"]["high_cited"],
        ]
    )
    ws.append(
        [
            "热点论文",
            summary["distribution"]["hot"]["1"],
            summary["distribution"]["hot"]["2"],
            summary["distribution"]["hot"]["3"],
            summary["distribution"]["hot"]["4"],
            summary["distribution"]["hot"]["5"],
            summary["distribution"]["hot"]["6"],
            summary["totals"]["hot"],
        ]
    )
    apply_header(ws, 1, 1, 8)
    apply_table_style(ws, 1, 3, 1, 8)
    autofit_like(ws, {"A": 16, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 10})


def build_subject_sheet(wb: Workbook, summary: dict) -> None:
    ws = wb.create_sheet("表13_院系学科")
    subjects = summary["unique_subjects"]
    headers = ["学院/学科"] + subjects + ["总计"]
    ws.append(headers)
    for college in summary["unique_colleges"]:
        row = [college]
        total = 0
        for subject in subjects:
            count = int(summary["college_subject_counts"].get(college, {}).get(subject, 0))
            row.append(count if count else "")
            total += count
        row.append(total)
        ws.append(row)
    apply_header(ws, 1, 1, len(headers))
    apply_table_style(ws, 1, ws.max_row, 1, len(headers))
    widths = {"A": 18}
    for idx in range(2, len(headers) + 1):
        widths[ws.cell(1, idx).column_letter] = 10
    autofit_like(ws, widths)


def build_author_sheet(wb: Workbook, summary: dict) -> None:
    ws = wb.create_sheet("表14_院系作者")
    ws.append(["学院", "作者", "篇数", "姓名状态"])
    pending_lookup = {}
    for item in summary.get("pinyin_pending_papers", []):
        for name in item.get("display_authors", []):
            pending_lookup[name] = "拼音待核"
    for college in summary["unique_colleges"]:
        author_counts = summary["author_counts"].get(college, {})
        for author, count in sorted(author_counts.items(), key=lambda kv: (-kv[1], kv[0])):
            status = pending_lookup.get(author, "已统一中文名")
            ws.append([college, author, count, status])
    apply_header(ws, 1, 1, 4)
    apply_table_style(ws, 1, ws.max_row, 1, 4)
    autofit_like(ws, {"A": 18, "B": 24, "C": 10, "D": 14})


def _paper_pending_map(summary: dict) -> dict[str, list[str]]:
    return {item["title"]: item.get("display_authors", []) for item in summary.get("pinyin_pending_papers", [])}


def build_papers_sheet(wb: Workbook, papers: list[dict], summary: dict, sheet_name: str = "去重总表") -> None:
    ws = wb.active
    ws.title = sheet_name
    pending_map = _paper_pending_map(summary)
    ws.append(
        [
            "序号",
            "论文题名",
            "高水平论文类型",
            "是否高被引",
            "高被引入围期数",
            "高被引入围期次",
            "是否热点",
            "热点入围期数",
            "热点入围期次",
            "所属学院",
            "所属ESI学科",
            "本校作者",
            "作者姓名状态",
            "原始作者串",
            "第一/通讯作者标记",
            "入围记录数",
            "入围原始记录",
        ]
    )

    def join_issues(items: list[int]) -> str:
        return "、".join(f"第{i}期" for i in items)

    def paper_type(item: dict) -> str:
        if item["high_cited_issues"] and item["hot_issues"]:
            return "高被引+热点"
        if item["high_cited_issues"]:
            return "高被引"
        return "热点"

    sorted_papers = sorted(
        papers,
        key=lambda item: (
            {"高被引+热点": 0, "高被引": 1, "热点": 2}[paper_type(item)],
            -len(item["high_cited_issues"]),
            -len(item["hot_issues"]),
            item["title"],
        ),
    )

    for index, paper in enumerate(sorted_papers, start=1):
        pending = pending_map.get(paper["title"], [])
        all_authors = list(paper["authors"]) + list(pending)
        if pending and paper["authors"]:
            status = "部分含拼音待核"
        elif pending:
            status = "含拼音待核"
        else:
            status = "已统一中文名"
        raw_author_strings = dedupe_preserve(record["authors_raw"] for record in paper.get("records", []))
        first_corr_flags = dedupe_preserve(record["first_corr"] for record in paper.get("records", []) if record.get("first_corr"))
        record_summaries = []
        for record in paper.get("records", []):
            piece = f"第{record['issue']}期-{record['category']}"
            if record.get("publication"):
                piece += f"-{record['publication']}"
            if record.get("year"):
                piece += f"-{record['year']}"
            if record.get("citations"):
                piece += f"-被引{record['citations']}"
            record_summaries.append(piece)
        ws.append(
            [
                index,
                paper["title"],
                paper_type(paper),
                "是" if paper["high_cited_issues"] else "否",
                len(paper["high_cited_issues"]),
                join_issues(paper["high_cited_issues"]),
                "是" if paper["hot_issues"] else "否",
                len(paper["hot_issues"]),
                join_issues(paper["hot_issues"]),
                "；".join(paper["colleges"]),
                "；".join(paper["subjects"]),
                "；".join(all_authors),
                status,
                " | ".join(raw_author_strings),
                "；".join(first_corr_flags),
                len(paper.get("records", [])),
                "；".join(record_summaries),
            ]
        )
    apply_header(ws, 1, 1, 17)
    apply_table_style(ws, 1, ws.max_row, 1, 17)
    autofit_like(
        ws,
        {
            "A": 8,
            "B": 70,
            "C": 14,
            "D": 10,
            "E": 12,
            "F": 18,
            "G": 10,
            "H": 12,
            "I": 18,
            "J": 24,
            "K": 18,
            "L": 28,
            "M": 14,
            "N": 36,
            "O": 14,
            "P": 12,
            "Q": 54,
        },
    )


def build_pending_sheet(wb: Workbook, summary: dict) -> None:
    ws = wb.create_sheet("拼音待核明细")
    ws.append(["论文题名", "已识别中文作者", "待核拼音作者", "所属学院", "高被引入围期次", "热点入围期次"])
    for item in summary.get("pinyin_pending_papers", []):
        ws.append(
            [
                item["title"],
                "；".join(item.get("resolved_authors", [])),
                "；".join(item.get("display_authors", [])),
                "；".join(item.get("colleges", [])),
                "、".join(f"第{i}期" for i in item.get("issues", {}).get("high_cited", [])),
                "、".join(f"第{i}期" for i in item.get("issues", {}).get("hot", [])),
            ]
        )
    apply_header(ws, 1, 1, 6, fill="C55A11")
    apply_table_style(ws, 1, ws.max_row, 1, 6)
    autofit_like(ws, {"A": 70, "B": 22, "C": 24, "D": 18, "E": 18, "F": 18})


def build_overview_sheet(wb: Workbook, summary: dict) -> None:
    ws = wb.create_sheet("统计汇总")
    ws.append(["2025年度ESI高水平论文去重汇总"])
    ws.append(["指标", "数值", "说明"])
    ws.append(["高水平论文总数", summary["totals"]["high_level"], "高被引论文与热点论文去重并集"])
    ws.append(["高被引论文", summary["totals"]["high_cited"], "去重后年度高被引论文数"])
    ws.append(["热点论文", summary["totals"]["hot"], "去重后年度热点论文数"])
    ws.append(["高被引与热点重合", summary["totals"]["high_cited"] + summary["totals"]["hot"] - summary["totals"]["high_level"], "同时属于高被引与热点的论文数"])
    ws.append([])
    ws.append(["兼容性提醒", len(summary.get("compatibility_warnings", [])), "上传文件名、sheet名、表头、模板结构与既有规则不完全一致时，会在详细总表中列出识别说明"])
    apply_header(ws, 1, 1, 3, fill="1F4E78")
    apply_header(ws, 2, 1, 3, fill="4F81BD")
    apply_table_style(ws, 1, ws.max_row, 1, 3)
    autofit_like(ws, {"A": 22, "B": 12, "C": 46})


def build_raw_records_sheet(wb: Workbook, papers: list[dict]) -> None:
    ws = wb.create_sheet("逐期原始记录")
    ws.append(
        [
            "序号",
            "论文题名",
            "期次",
            "名单类型",
            "出版物",
            "被引次数",
            "出版时间",
            "第一/通讯作者",
            "原始作者串",
            "所属学院",
            "所属ESI学科",
        ]
    )
    row_no = 1
    for paper in papers:
        for record in paper.get("records", []):
            ws.append(
                [
                    row_no,
                    paper["title"],
                    f"第{record['issue']}期",
                    record["category"],
                    record.get("publication", ""),
                    record.get("citations", ""),
                    record.get("year", ""),
                    record.get("first_corr", ""),
                    record.get("authors_raw", ""),
                    "；".join(record.get("colleges", [])),
                    "；".join(record.get("subjects", [])),
                ]
            )
            row_no += 1
    apply_header(ws, 1, 1, 11, fill="5B9BD5")
    apply_table_style(ws, 1, ws.max_row, 1, 11)
    autofit_like(ws, {"A": 8, "B": 60, "C": 10, "D": 12, "E": 22, "F": 10, "G": 12, "H": 14, "I": 32, "J": 18, "K": 18})


def build_compatibility_sheet(wb: Workbook, summary: dict) -> None:
    ws = wb.create_sheet("兼容性检查")
    ws.append(["检查项", "结果", "说明"])
    warnings = summary.get("compatibility_warnings", [])
    ws.append(["整体状态", "存在兼容性提示" if warnings else "正常", "有提示不代表结果错误，但建议核对识别说明"])
    for item in summary.get("source_resolution", []):
        result = "自动匹配" if item.get("warning") else "精确匹配"
        detail = f"文件：{Path(item['file']).name}；预期sheet：{item['preferred_sheet']}；实际sheet：{item['resolved_sheet']}；记录数：{item['record_count']}"
        if item.get("warning"):
            detail += f"；提示：{item['warning']}"
        ws.append([f"第{item['issue']}期附表", result, detail])
    for warning in warnings:
        ws.append(["兼容性提示", "请复核", warning])
    apply_header(ws, 1, 1, 3, fill="C55A11")
    apply_table_style(ws, 1, ws.max_row, 1, 3)
    autofit_like(ws, {"A": 20, "B": 12, "C": 84})


def generate_stats_workbook(analysis_dir: Path, output_xlsx: Path) -> Path:
    summary = load_json(analysis_dir / "summary.json")

    wb = Workbook()
    build_distribution_sheet(wb, summary)
    build_subject_sheet(wb, summary)
    build_author_sheet(wb, summary)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    return output_xlsx


def generate_detailed_workbook(analysis_dir: Path, output_xlsx: Path) -> Path:
    summary = load_json(analysis_dir / "summary.json")
    papers = load_json(analysis_dir / "papers.json")

    wb = Workbook()
    build_papers_sheet(wb, papers, summary)
    build_overview_sheet(wb, summary)
    build_pending_sheet(wb, summary)
    build_raw_records_sheet(wb, papers)
    build_compatibility_sheet(wb, summary)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    return output_xlsx


def main() -> None:
    args = parse_args()
    stats_output = generate_stats_workbook(args.analysis_dir, args.output_xlsx)
    detailed_output = generate_detailed_workbook(args.analysis_dir, args.detailed_output_xlsx)
    print(stats_output)
    print(detailed_output)


if __name__ == "__main__":
    main()
